import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():

    # API
    Api_url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open('C:\\Users\\rkusuma\\Desktop\\API\\addstudent2.json', 'r')
    json_request = json.loads(file.read())

    # Excel code
    # we need to open workbook here openg workbook is python concept
    vk = openpyxl.load_workbook('C:\\Users\\rkusuma\\Desktop\\API\\Test_data_exl.xlsx')
    #now i need to move to sheet am working
    sh = vk['Sheet1']
    #now i need to find out how many test data it has in sheet
    #it fetches max rows i have in exl
    rows = sh.max_row

    #now am running a loop for each row
    # as 1 st row is header am giving from 2 and also rows+1 because it skips last row so we need to use +1

    print(rows)

    for i in range(2,rows+1):

        cell_first_name = sh.cell( row=i, column=1)
        print(cell_first_name)
        print(cell_first_name.value)
        json_request['first_name'] = cell_first_name.value

        cell_mid_name = sh.cell(row=i, column=2)
        print(cell_mid_name)
        print(cell_mid_name.value)
        json_request['middle_name'] = cell_mid_name.value

        cell_last_name = sh.cell(row=i, column=3)
        print(cell_last_name)
        print(cell_last_name.value)
        json_request['last_name'] = cell_last_name.value

        cell_dob = sh.cell(row=i, column=4)
        print(cell_dob)
        print(cell_dob.value)
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(Api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201



