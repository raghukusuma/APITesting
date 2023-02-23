import requests
import json
import jsonpath
import openpyxl

#some errors its showing here so see here fr explanation and for code check always_data_driven_working_code directry

def test_add_multiple_students():

    # API
    Api_url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open('C:\\Users\\rkusuma\\Desktop\\API\\addstudent2.json', 'r')
    json_request = json.loads(file.read())

    # Excel code
    # we need to open workbook here openg workbook is python concept
    vk = openpyxl.load_workbook('C:\\Users\\rkusuma\\Desktop\\API\\Test_data_exl.xlsx')
    #now i need to move to sheet am working
    sh = vk['Sheet1']
    #now i need to find out how many test data it has in sheet
    #it fetches max rows i have in exl
    rows = sh.max_row

    #now am running a loop for each row
    # as 1 st row is header am giving from 2 and also rows+1 because it skips last row so we need to use +1


    for i in range(2,rows+1):

        # first_name in evry row is first column
        cell_first_name = sh.cell( row=i, column=1)
        # mid_name in evry row is 2nd column
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        #we are creating 4 ojects fr each cell in evry row
        #read data from cell
        # now i want to put these values in json

         json_request['first_name']=cell_first_name.value
         json_request['middle_name']=cell_mid_name.value
         json_request['last_name']=cell_last_name.value
         json_request['date_of_birth']=cell_dob.value

#now we got values from excl and then we posted them into values of json now we can post them onto url

         response = requests.post(Api_url, json_request)
         print(response.text)
         print(response.status_code)
         assert response.status_code==201