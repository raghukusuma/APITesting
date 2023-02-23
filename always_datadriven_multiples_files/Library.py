
# our code to multiple files so dynamic - and applicaple to any test cae

# library file and test case file we are dng

#create a class n inside we are gng to use many methods


import json
import jsonpath
import requests
import openpyxl

#we are gng to create few methods now here

#create method to find row count

class common:



    #method name
   # def fetch_row_count(self, FileNamePath, SheetName ):
    #vk = openpyxl.load_workbook('C:\\Users\\rkusuma\\Desktop\\API\\Test_data_exl.xlsx') like always_datadriven_working_code directry code
    #we dnt use static path now but we use what evr path we pass to abve argument FileNamePath
    #vk = openpyxl.load_workbook(FileNamePath)
    #sh = vk['Sheet1']
    #what evr sheet name we are giving we use it here now
    #sh = vk[SheetName]
   # rows = sh.max_row

    #method 1 is ready fetching no.of rows and returning it

    #method 2 to fetch no.of columns

    #def fetch_column_count(self, FileNamePath, SheetName ):
    # option 1 passing values again like previous function
    #option 2 , am creating a constructor here so i can use them at any wr .. also i need to create a global constructr


    # so am writing code again with constructrs


    def __init__(self, FileNamePath, SheetName):
        global vk
        global sh

        #sh is sheet objct and vk is workbook object
        #created belw values as gloabl in constrctrs so i can use them at any wr
        vk = openpyxl.load_workbook(FileNamePath)
        sh = vk[SheetName]


        def fetch_row_count(self):
            rows = sh.max_row
            return rows

        def fetch_column_count(self):
            col = sh.max_column
            return col

    #now we need to what all keys we are having in json

         def fetch_key_names(selfself):
             c=sh.max_column
             li=[]
             for i in range(1, c+1):
                 cell = sh.cell (row=1, column=i)
                 li.insert(i-1,cell.value)

             return li

        #now we neeed to create one mre method - to update json report with mutiple data

        def update_request_with_data(self,rowNumber,jsonRequest,keyList):

            c= sh.max_column
            for i in range (1,c+1):
                    cell = sh.cell(row=rowNumber, column=i)
                    jsonRequest[keyList[i-1]]=cell.value

            return jsonRequest

        #library is prepared now we are ready to write testcases in anthr file which we created
